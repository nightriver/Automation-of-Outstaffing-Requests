"""
Скрипт підготовки шаблону assets/client_template.xlsx
"""

import os
from openpyxl import Workbook, load_workbook
from core.schema import TEMPLATE_HEADERS, validate_template_headers


def main():
    os.makedirs("assets", exist_ok=True)
    source_path = os.path.join("docs", "Приклад таблички от клієнта.xlsx")
    target_path = os.path.join("assets", "client_template.xlsx")

    if os.path.exists(source_path):
        try:
            wb = load_workbook(source_path)
            ws = wb.active
            if ws.title != "Реєстр_Даних":
                ws.title = "Реєстр_Даних"
            # Записуємо обов'язкові 27 заголовків у рядку 1
            for idx, header in enumerate(TEMPLATE_HEADERS, start=1):
                ws.cell(row=1, column=idx, value=header)
            validate_template_headers(ws)
            wb.save(target_path)
            print(f"Шаблон успішно створено з {source_path} -> {target_path}")
            return
        except Exception as e:
            print(f"Помилка відкриття вихідного файлу: {e}")

    # Фолбек: створення нового чистого шаблону з нуля
    wb = Workbook()
    ws = wb.active
    ws.title = "Реєстр_Даних"
    for idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)
    validate_template_headers(ws)
    wb.save(target_path)
    print(f"Створено новий базований шаблон -> {target_path}")


if __name__ == "__main__":
    main()
