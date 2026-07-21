"""
Модуль рендеру друкованого листа 'Форма_Печати' відповідно до офіційного договору-заявки
З підстановкою назви Замовника та Директора Замовника у преамбулу та підписи.
Знято захист листа для вільного редагування менеджером.
"""

import datetime
import math
from openpyxl.styles import Alignment, Font, PatternFill

YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
CHARS_PER_LINE = 85       # Розрахункова ширина текстової зони колонок A:C
LINE_HEIGHT_PT = 16       # Висота одного рядка тексту в пунктах
MIN_ROW_HEIGHT_PT = 20    # Мінімальна базова висота рядка бланка


def set_merged_row_height(ws, row: int, text: str) -> None:
    """
    Розраховує та встановлює пропорційну висоту об'єднаного рядка
    із коефіцієнтом запасу 1.30 для запобігання обрізанню нижніх елементів літер у PDF.
    """
    if not text:
        ws.row_dimensions[row].height = MIN_ROW_HEIGHT_PT
        return

    explicit_paragraphs = str(text).replace("\r\n", "\n").split("\n")
    total_lines = 0
    for paragraph in explicit_paragraphs:
        paragraph_len = len(paragraph) if paragraph else 1
        total_lines += max(1, math.ceil(paragraph_len / CHARS_PER_LINE))

    ws.row_dimensions[row].height = max(
        MIN_ROW_HEIGHT_PT,
        math.ceil(total_lines * LINE_HEIGHT_PT * 1.30),
    )


def split_lines(value: str) -> list[str]:
    """
    Розбиває багаторядковий текст на список непорожніх рядків.
    """
    return [line.strip() for line in str(value or "").replace("\r\n", "\n").split("\n") if line.strip()]


def build_print_sheet(wb, ws_data, client_template: dict) -> None:
    """
    Створює та форматує друкований лист 'Форма_Печати' відповідно до структури Word-документа.
    """
    ws_p = wb.create_sheet("Форма_Печати")
    ws_p.views.sheetView[0].showGridLines = False

    # Налаштування колонок
    ws_p.column_dimensions["A"].width = 14
    ws_p.column_dimensions["B"].width = 45
    ws_p.column_dimensions["C"].width = 35

    font_title = Font(name="Arial", size=12, bold=True)
    font_subtitle = Font(name="Arial", size=10, italic=True)
    font_section = Font(name="Arial", size=10, bold=True)
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # Отримання полів Замовника та Директора з client_template
    ct = client_template or {}
    client_name = ct.get("client_name") or "ТОВ «_________________»"
    client_director = ct.get("client_director") or "_________________"
    contract_info = ct.get("contract_info", "№ 2502-ЕК/ДЕС-26 від 25.02.2026")
    appendix_info = ct.get("appendix_info", "№ 1 / OF-26 від 25.02.2026")
    provider_name = ct.get("provider_name", "ТОВ «ДЖІ ЕС СТАФФІНГ»")
    provider_director = ct.get("provider_director", "Воронова І.С.")

    # 1. Шапка документа
    ws_p.merge_cells("A1:C1")
    ws_p["A1"] = "Заявка"
    ws_p["A1"].font = font_title
    ws_p["A1"].alignment = align_center

    ws_p.merge_cells("A2:C2")
    ws_p["A2"] = f"до Договору про надання послуг {contract_info}, далі іменований \"Договір\""
    ws_p["A2"].font = font_subtitle
    ws_p["A2"].alignment = align_center

    # Номер заявки (формула на A2 Листа 1)
    ws_p["A3"] = "Заявка №"
    ws_p["A3"].font = font_bold
    ws_p["B3"] = "='Реєстр_Даних'!A2"
    ws_p["B3"].alignment = align_left

    # Місто та Дата заявки (C4 розблокована жовта комірка)
    ws_p["A4"] = "м. Київ"
    ws_p["A4"].font = font_regular
    ws_p["C4"] = datetime.date.today().strftime("«%d» %m %Y року")
    ws_p["C4"].fill = YELLOW_FILL
    ws_p["C4"].alignment = align_right

    # Преамбула (з підстановкою назви Замовника)
    row_ptr = 6
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    preamble_1 = f"{client_name}, далі іменоване \"Замовник\", направляє до {provider_name}, далі іменоване \"Виконавець\", Заявку для розгляду можливості її виконання силами Виконавця."
    ws_p.cell(row=row_ptr, column=1, value=preamble_1).font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left
    set_merged_row_height(ws_p, row_ptr, preamble_1)

    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    preamble_2 = "Ця Заявка складена в рамках надання послуг із забезпечення Замовника персоналом у формі адміністрування та направлення Працівників у розпорядження Замовника, для виконання певних функцій, силами Виконавця за Заявкою Замовника на наступних умовах:"
    ws_p.cell(row=row_ptr, column=1, value=preamble_2).font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left
    set_merged_row_height(ws_p, row_ptr, preamble_2)

    # Розділ 1. Замовник
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value=f"1. Замовник: {client_name}").font = font_section

    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value=f"Відповідальна особа з боку Замовника: {client_director}").font = font_regular
    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value="Телефон відповідальної особи: —").font = font_regular
    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value="Електронна пошта відповідальної особи: —").font = font_regular

    # Розділ 2. Послуга
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2. Послуга").font = font_section

    # 2.1. Вид потрібної послуги
    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value="2.1. Вид потрібної послуги та її опис").font = font_bold
    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    cell_21 = ws_p.cell(row=row_ptr, column=1, value="='Реєстр_Даних'!G2")
    cell_21.font = font_regular
    cell_21.alignment = align_left

    # 2.2. Термін надання послуги
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.2. Термін надання послуги").font = font_bold
    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    ws_p.cell(row=row_ptr, column=1, value="='Реєстр_Даних'!E2").font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left

    term_desc = "на період до завершення проєкту та отримання від Замовника повідомлення про припинення послуги або зміну її обсягу."
    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    ws_p.cell(row=row_ptr, column=1, value=term_desc).font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left
    set_merged_row_height(ws_p, row_ptr, term_desc)

    # 2.3. Місце надання послуги
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.3. Місце надання послуги").font = font_bold
    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    cell_23 = ws_p.cell(row=row_ptr, column=1, value="='Реєстр_Даних'!S2")
    cell_23.font = font_regular
    cell_23.alignment = align_left

    # 2.4. Час надання послуги
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.4. Час надання послуги").font = font_bold
    row_ptr += 1
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    ws_p.cell(row=row_ptr, column=1, value="='Реєстр_Даних'!Q2").font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left

    # 2.5. Вимоги до кваліфікації Працівників
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.5. Вимоги до кваліфікації Працівників").font = font_bold
    from core.schema import HEADER_TO_COL
    raw_notes = ws_data.cell(row=2, column=HEADER_TO_COL["Примітка"]).value or ""
    notes_lines = split_lines(raw_notes)
    if not notes_lines:
        notes_lines = ["Вища або середня спеціальна освіта, відповідність професійним вимогам."]

    for note in notes_lines:
        row_ptr += 1
        ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
        cell = ws_p.cell(row=row_ptr, column=1, value=f"• {note}")
        cell.font = font_regular
        cell.alignment = align_left
        set_merged_row_height(ws_p, row_ptr, note)

    # 2.6. Завдання, які можуть надаватися Працівникам
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.6. Завдання, які можуть надаватися Працівникам").font = font_bold
    raw_duties = ws_data.cell(row=2, column=HEADER_TO_COL["Функціональні обов'язки"]).value or ""
    duties_lines = split_lines(raw_duties)
    if not duties_lines:
        duties_lines = ["Виконання функціональних обов'язків згідно з інструкцією."]

    for idx, duty in enumerate(duties_lines, start=1):
        row_ptr += 1
        ws_p.cell(row=row_ptr, column=1, value=f"{idx}.").alignment = align_center
        ws_p.cell(row=row_ptr, column=1).font = font_bold

        ws_p.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=3)
        cell = ws_p.cell(row=row_ptr, column=2, value=duty)
        cell.font = font_regular
        cell.alignment = align_left
        set_merged_row_height(ws_p, row_ptr, duty)

    # 2.7. Додаткові вимоги
    row_ptr += 2
    ws_p.cell(row=row_ptr, column=1, value="2.7. Додаткові вимоги – немає.").font = font_bold

    # Розділ 3. Заключні положення
    row_ptr += 2
    ws_p.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
    final_text = f"3. Ця Заявка є невід'ємною частиною Додатку {appendix_info} до Договору про надання послуг {contract_info}."
    ws_p.cell(row=row_ptr, column=1, value=final_text).font = font_regular
    ws_p.cell(row=row_ptr, column=1).alignment = align_left
    set_merged_row_height(ws_p, row_ptr, final_text)

    # Блок підписів в кінці документа
    row_ptr += 3
    ws_p.cell(row=row_ptr, column=1, value="ЗАМОВНИК").font = font_bold
    ws_p.cell(row=row_ptr, column=3, value="ВИКОНАВЕЦЬ").font = font_bold

    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value=client_name).font = font_bold
    ws_p.cell(row=row_ptr, column=3, value=provider_name).font = font_bold

    row_ptr += 1
    ws_p.cell(row=row_ptr, column=1, value=f"Директор ___________ / {client_director} /").font = font_regular
    ws_p.cell(row=row_ptr, column=3, value=f"Директор ___________ / {provider_director} /").font = font_regular

    # ЗАХИСТ ЗНЯТО: Лист повністю редагований менеджером у Excel/LibreOffice
    ws_p.protection.sheet = False

    # Налаштування області друку для формату А4
    ws_p.page_setup.orientation = "portrait"
    ws_p.page_setup.paperSize = ws_p.PAPERSIZE_A4
    ws_p.sheet_properties.pageSetUpPr.fitToPage = True
    ws_p.page_setup.fitToWidth = 1
    ws_p.page_setup.fitToHeight = 0
    ws_p.print_area = f"A1:C{max(row_ptr, 35)}"
