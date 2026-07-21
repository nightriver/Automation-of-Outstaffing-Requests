"""
Головний диспетчер генерації Excel-пакета в пам'яті (BytesIO).
"""

from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from core.export_instruction import build_instruction_sheet
from core.export_print import build_print_sheet
from core.schema import HEADER_TO_COL, TEMPLATE_HEADERS, validate_template_headers
from core.security import write_client_value

YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")


def create_base_template_workbook() -> Workbook:
    """
    Створює базову робочу книгу з 27 заголовками у Ррядку 1 Листа 1 ('Реєстр_Даних').
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Реєстр_Даних"
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    return wb


def build_workbook(template_path: str, form_data: dict, client_template=None) -> bytes:
    """
    Приймає шлях до шаблону, наповнює дані форми у Рядок 2,
    передає реквізити Замовника та Директора на Лист 2 і повертає масив байтів.
    """
    if template_path:
        try:
            wb = load_workbook(template_path)
        except Exception:
            wb = create_base_template_workbook()
    else:
        wb = create_base_template_workbook()

    ws = wb[wb.sheetnames[0]]
    if ws.title != "Реєстр_Даних":
        ws.title = "Реєстр_Даних"

    validate_template_headers(ws)

    # Запис клієнтських даних у Рядок 2
    for header, value in form_data.items():
        if header not in HEADER_TO_COL:
            continue
        col = HEADER_TO_COL[header]
        write_client_value(ws, 2, col, value)

    # Оформлення службової комірки А2 для менеджера
    ws["A2"].value = None
    ws["A2"].fill = YELLOW_FILL
    ws["A2"].comment = Comment(
        "Поле заповнює менеджер Smart Solutions після отримання заявки",
        "system",
    )

    # Підготовка метаданих client_template для Листа 2
    if isinstance(client_template, dict):
        ct = dict(client_template)
    else:
        ct = {}

    if "Замовник" in form_data and form_data["Замовник"]:
        ct["client_name"] = form_data["Замовник"]
    if "Директор замовника" in form_data and form_data["Директор замовника"]:
        ct["client_director"] = form_data["Директор замовника"]

    # Збирання суміжних листів
    build_print_sheet(wb, ws, ct)
    build_instruction_sheet(wb)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
