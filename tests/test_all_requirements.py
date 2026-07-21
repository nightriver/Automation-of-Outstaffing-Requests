"""
Автоматичний тестовий комплект pytest для перевірки 9 обов'язкових вимого з Розділу 15 ТЗ v4.2.
"""

from io import BytesIO
from openpyxl import load_workbook
import pytest

from core.export import build_workbook
from core.export_print import set_merged_row_height
from core.mailer import MockMailProvider, OutstaffingMailer
from core.routes import resolve_route
from core.schema import TEMPLATE_HEADERS
from core.security import escape_excel_text


@pytest.fixture
def sample_form_data():
    return {
        "Замовник": "ТОВ «ОБРІЙ - 2020»",
        "Директор замовника": "Чекман П.С.",
        "П.І.Б.": "Бондар Денис Миколайович",
        "ІПН": "1234567890",
        "Дата прийома": "01.09.2026",
        "Термін дії СТД": "до завершення Заявки",
        "Посада згідно кп": "Менеджер з продажів",
        "Випробувальний термін": "3 місяці",
        "Номер телефону": "+380991234567",
        "E-mail": "bondar@example.com",
        "Оклад гросс": "35000",
        "Бонус": "10000",
        "Періодичність бонуса": "Щомісячно",
        "Робота": "Повна зайнятість",
        "Дата виплата авансу та ЗП": "15 та 30 числа",
        "Графік роботи (з-по)": "09:00 - 18:00",
        "Графік роботи количество годин": "40",
        "Кількість дней відпустки": "24",
        "Місто та місце роботи": "м. Київ, вул. Хрещатик, 1",
        "Формат роботи": "Офіс",
        "Матеріальна відповідальність": "Ні",
        "Інші компенсації": "Мобільний зв'язок",
        "Коментарі": "Без зауважень",
        "Реквізити картки": "UA123456789012345678901234567",
        "Функціональні обов'язки": "=SUM(1,2)\nПроведення переговорів\nПідготовка звітів",
        "Примітка": "Категорія B",
        "документи": "Паспорт, ІПН",
    }


# 1. Тест суворого контракту заголовків
def test_1_strict_headers_contract(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb["Реєстр_Даних"]

    actual_headers = [ws.cell(1, i).value for i in range(1, len(TEMPLATE_HEADERS) + 1)]
    assert actual_headers == TEMPLATE_HEADERS
    assert ws.max_column == 27


# 2. Тест цілісності комірки A2
def test_2_cell_a2_integrity(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb["Реєстр_Даних"]

    assert ws["A2"].value is None
    assert ws["A2"].fill.fgColor.rgb in ["00FFF2CC", "FFF2CC"]
    assert ws["A2"].comment is not None
    assert "менеджер Smart Solutions" in ws["A2"].comment.text


# 3. Тест очищення префіксів (Excel Injection)
def test_3_excel_injection_sanitization():
    assert escape_excel_text("=1+1") == "'=1+1"
    assert escape_excel_text("+38099123") == "'+38099123"
    assert escape_excel_text("-100") == "'-100"
    assert escape_excel_text("@cmd") == "'@cmd"
    assert escape_excel_text("Звичайний текст") == "Звичайний текст"


# 4. Тест формул друкованої форми (Заявка № -> B3)
def test_4_print_sheet_formulas(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    wb = load_workbook(BytesIO(file_bytes))
    ws_p = wb["Форма_Печати"]

    assert ws_p["B3"].value == "='Реєстр_Даних'!A2"


# 5. Тест автоматичного розрахунку висоти рядків
def test_5_row_height_calculation():
    wb = load_workbook("assets/client_template.xlsx")
    ws = wb.active

    long_text = "Довгий текст обов'язку " * 10
    set_merged_row_height(ws, 20, long_text)
    assert ws.row_dimensions[20].height > 20


# 6. Тест зняття захисту з 2 листа
def test_6_sheet_protection_unlocked(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    wb = load_workbook(BytesIO(file_bytes))
    ws_p = wb["Форма_Печати"]

    assert ws_p.protection.sheet is False


# 7. Тест успішного сценарію надсилання
def test_7_mailer_success_scenario(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    mock_provider = MockMailProvider(should_succeed=True)
    mailer = OutstaffingMailer(mock_provider, "notifications@smart-solutions.ua")

    ok = mailer.send_document_package(file_bytes, "test.xlsx", "manager@smart-hr.com.ua", "SS-20260720-0001")
    assert ok is True
    assert len(mock_provider.sent_messages) == 1


# 8. Тест fallback-сценарію
def test_8_mailer_fallback_scenario(sample_form_data):
    file_bytes = build_workbook("assets/client_template.xlsx", sample_form_data)
    mock_provider = MockMailProvider(should_succeed=False)
    mailer = OutstaffingMailer(mock_provider, "notifications@smart-solutions.ua")

    ok = mailer.send_document_package(file_bytes, "test.xlsx", "manager@smart-hr.com.ua", "SS-20260720-0001")
    assert ok is False


# 9. Тест ізоляції secrets / route key
def test_9_secrets_route_key_isolation(monkeypatch):
    import streamlit as st

    fake_secrets = {
        "routes": {
            "valid_key": {
                "route_id": "R001",
                "active": True,
                "email_secret_key": "mgr1",
                "client_template": "std",
            },
            "inactive_key": {
                "route_id": "R002",
                "active": False,
                "email_secret_key": "mgr1",
                "client_template": "std",
            },
        },
        "managers": {"mgr1": "manager@smart-hr.com.ua"},
    }
    monkeypatch.setattr(st, "secrets", fake_secrets)

    resolved = resolve_route("valid_key")
    assert resolved["email"] == "manager@smart-hr.com.ua"

    with pytest.raises(ValueError):
        resolve_route("inactive_key")

    with pytest.raises(ValueError):
        resolve_route("unknown_key")
